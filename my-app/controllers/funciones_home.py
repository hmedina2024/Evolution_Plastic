
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file

### Empleados
def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando documento
    documento_sin_puntos = re.sub('[^0-9]+', '', dataForm['documento'])
    # convertir documento a INT
    documento = int(documento_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (documento,nombre_empleado, apellido_empleado, tipo_empleado, telefono_empleado, email_empleado, cargo, foto_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (documento,dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['tipo_empleado'],
                            dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['cargo'], result_foto_perfil)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.documento,
                        e.nombre_empleado, 
                        e.apellido_empleado,                        
                        e.foto_empleado,
                        e.cargo,
                        CASE
                            WHEN e.tipo_empleado = 1 THEN 'Directo'
                            ELSE 'Temporal'
                        END AS tipo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
                print(empleadosBD)
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.documento,
                        e.nombre_empleado, 
                        e.apellido_empleado,                        
                        CASE
                            WHEN e.tipo_empleado = 1 THEN 'Directo'
                            ELSE 'Temporal'
                        END AS tipo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.cargo,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.documento,
                        e.nombre_empleado, 
                        e.apellido_empleado,                        
                        e.email_empleado,
                        e.telefono_empleado,
                        e.cargo,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.tipo_empleado = 1 THEN 'Directo'
                            ELSE 'Temporal'
                        END AS tipo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Documento","Nombre", "Apellido", "Tipo Empleado",
                     "Telefono", "Email", "Profesión", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        documento = registro['documento']
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        tipo_empleado = registro['tipo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        cargo = registro['cargo']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((documento,nombre_empleado, apellido_empleado, tipo_empleado, telefono_empleado, email_empleado, cargo,
                      fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.documento,
                            e.nombre_empleado, 
                            e.apellido_empleado,                            
                            CASE
                                WHEN e.tipo_empleado = 1 THEN 'Directo'
                                ELSE 'Temporal'
                            END AS tipo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s  
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.documento,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.tipo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.cargo,                            
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                tipo_empleado = data.form['tipo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                cargo = data.form['cargo']

                documento_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['documento'])
                documento = int(documento_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            documento = %s,
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            tipo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            cargo = %s,                            
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (documento,nombre_empleado, apellido_empleado, tipo_empleado,
                                telefono_empleado, email_empleado, cargo,
                                fotoForm, id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            documento = %s,
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            tipo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            cargo = %s                            
                        WHERE id_empleado = %s
                    """
                    values = (documento, nombre_empleado, apellido_empleado, tipo_empleado,
                                telefono_empleado, email_empleado, cargo,
                                id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar Empleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []







### PROCESOS    
def procesar_form_proceso(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_procesos (codigo_proceso, nombre_proceso, descripcion_proceso) VALUES (%s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['cod_proceso'], dataForm['nombre_proceso'], dataForm['descripcion_proceso'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_proceso: {str(e)}'


# Lista de Procesos
def sql_lista_procesosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        p.id_proceso,
                        p.codigo_proceso,
                        p.nombre_proceso,
                        p.descripcion_proceso,                        
                        p.fecha_registro
                    FROM tbl_procesos AS p
                    ORDER BY p.id_proceso DESC
                    """
                cursor.execute(querySQL)
                procesosBD = cursor.fetchall()
        return procesosBD
    except Exception as e:
        print(f"Error en la función sql_lista_procesosBD: {e}")
        return None


# Detalles del Proceso
def sql_detalles_procesosBD(id_proceso):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        p.id_proceso,
                        p.codigo_proceso,
                        p.nombre_proceso,
                        p.descripcion_proceso,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_procesos AS p
                    WHERE codigo_proceso =%s
                    ORDER BY p.id_proceso DESC
                    """)
                cursor.execute(querySQL, (id_proceso,))
                procesosBD = cursor.fetchone()
        return procesosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None



def buscarProcesoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            p.id_proceso,
                            p.codigo_proceso,
                            p.nombre_proceso,
                            p.descripcion_proceso,                        
                            p.fecha_registro
                        FROM tbl_procesos AS p
                        WHERE p.id_proceso =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                proceso = mycursor.fetchone()
                return proceso

    except Exception as e:
        print(f"Ocurrió un error en def buscarProcesoUnico: {e}")
        return []


def procesar_actualizar_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                codigo_proceso = data.form['codigo_proceso']
                nombre_proceso = data.form['nombre_proceso']
                descripcion_proceso = data.form['descripcion_proceso']
                id_proceso = data.form['id_proceso']             
                querySQL = """
                    UPDATE tbl_procesos
                    SET 
                        codigo_proceso = %s,
                        nombre_proceso = %s,
                        descripcion_proceso = %s
                    WHERE id_proceso = %s
                """
                values = (codigo_proceso, nombre_proceso, descripcion_proceso,id_proceso)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizar_form: {e}")
        return None

# Eliminar Procesos
def eliminarProceso(id_proceso):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_procesos WHERE id_proceso=%s"
                cursor.execute(querySQL, (id_proceso,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarProceso : {e}")
        return []
    
    
    
    


### CLIENTES    
def procesar_form_cliente(dataForm, foto_perfil_cliente):
    # Formateando documento
    documento_sin_puntos = re.sub('[^0-9]+', '', dataForm['documento'])
    # convertir documento a INT
    documento = int(documento_sin_puntos)

    result_foto_cliente = procesar_imagen_cliente(foto_perfil_cliente)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_clientes (tipo_documento,documento,nombre_cliente, telefono_cliente, email_cliente, foto_cliente) VALUES (%s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['tipo_documento'],documento,dataForm['nombre_cliente'],dataForm['telefono_cliente'], dataForm['email_cliente'], result_foto_cliente)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_cliente: {str(e)}'


def procesar_imagen_cliente(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_clientes/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Clientes
def sql_lista_clientesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_cliente,                        
                        e.tipo_documento,
                        e.documento,
                        e.nombre_cliente, 
                        e.telefono_cliente,                        
                        e.foto_cliente,
                        e.email_cliente                        
                    FROM tbl_clientes AS e
                    ORDER BY e.id_cliente DESC
                    """)
                cursor.execute(querySQL,)
                clientesBD = cursor.fetchall()
                print(clientesBD)
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_clientesBD: {e}")
        return None


# Detalles del Cliente
def sql_detalles_clientesBD(idCliente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_cliente,
                        e.tipo_documento,
                        e.documento,
                        e.nombre_cliente,                      
                        e.telefono_cliente, 
                        e.email_cliente,
                        e.foto_cliente,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_clientes AS e
                    WHERE id_cliente =%s
                    ORDER BY e.id_cliente DESC
                    """)
                cursor.execute(querySQL, (idCliente,))
                clientesBD = cursor.fetchone()
        return clientesBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_clientesBD: {e}")
        return None


def buscarClienteBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_cliente,
                            e.tipo_documento,
                            e.documento,
                            e.nombre_cliente, 
                            e.email_cliente,            
                        FROM tbl_clientes AS e
                        WHERE e.nombre_cliente LIKE %s  
                        ORDER BY e.id_cliente DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarClienteBD: {e}")
        return []


def buscarClienteUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_cliente,
                            e.documento,
                            e.nombre_cliente, 
                            e.tipo_documento,
                            e.telefono_cliente,
                            e.email_cliente,
                            e.foto_cliente
                        FROM tbl_clientes AS e
                        WHERE e.id_cliente =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                cliente = mycursor.fetchone()
                return cliente

    except Exception as e:
        print(f"Ocurrió un error en def buscarClienteUnico: {e}")
        return []


def procesar_actualizacion_cliente(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                tipo_documento = data.form['tipo_documento']
                nombre_cliente = data.form['nombre_cliente']               
                telefono_cliente = data.form['telefono_cliente']
                email_cliente = data.form['email_cliente']
                documento_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['documento'])
                documento = int(documento_sin_puntos)
                id_cliente = data.form['id_cliente']
                if data.files['foto_cliente']:
                    file = data.files['foto_cliente']
                    fotoForm = procesar_imagen_cliente(file)
                    querySQL = """
                        UPDATE tbl_clientes
                        SET 
                            tipo_documento = %s,
                            nombre_cliente = %s,                                                       
                            telefono_cliente = %s,
                            email_cliente = %s,
                            documento = %s,                            
                            foto_cliente = %s
                        WHERE id_cliente = %s
                    """
                    values = (tipo_documento,nombre_cliente,telefono_cliente, email_cliente,documento,
                                fotoForm, id_cliente)
                else:
                    querySQL = """
                        UPDATE tbl_clientes
                        SET 
                            tipo_documento = %s,
                            nombre_cliente = %s,                                                       
                            telefono_cliente = %s,
                            email_cliente = %s,
                            documento = %s                            
                        WHERE id_cliente = %s
                    """
                    values = (tipo_documento,nombre_cliente,telefono_cliente, email_cliente,documento,id_cliente)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_cliente: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar Cliente
def eliminarCliente(id_cliente, foto_cliente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_clientes WHERE id_cliente=%s"
                cursor.execute(querySQL, (id_cliente,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_clientes', foto_cliente)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarCliente : {e}")
        return []













### ACTIVIDADES    
def procesar_form_actividad(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_actividades (codigo_actividad, nombre_actividad, descripcion_actividad) VALUES (%s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['cod_actividad'], dataForm['nombre_actividad'], dataForm['descripcion_actividad'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_actividad: {str(e)}'


# Lista de Actividades
def sql_lista_actividadesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        a.id_actividad,
                        a.codigo_actividad,
                        a.nombre_actividad,
                        a.descripcion_actividad,                        
                        a.fecha_registro
                    FROM tbl_actividades AS a
                    ORDER BY a.id_actividad DESC
                    """
                cursor.execute(querySQL)
                actividadesBD = cursor.fetchall()
        return actividadesBD
    except Exception as e:
        print(f"Error en la función sql_lista_actividadesBD: {e}")
        return None


# Detalles de la actividad
def sql_detalles_actividadesBD(id_actividad):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        a.id_actividad,
                        a.codigo_actividad,
                        a.nombre_actividad,
                        a.descripcion_actividad,
                        DATE_FORMAT(a.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_actividades AS a
                    WHERE codigo_actividad =%s
                    ORDER BY a.id_actividad DESC
                    """)
                cursor.execute(querySQL, (id_actividad,))
                actividadBD = cursor.fetchone()
        return actividadBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_actividadesBD: {e}")
        return None


def buscarActividadUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            a.id_actividad,
                            a.codigo_actividad,
                            a.nombre_actividad,
                            a.descripcion_actividad,                        
                            a.fecha_registro
                        FROM tbl_actividades AS a
                        WHERE a.id_actividad =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                actividad = mycursor.fetchone()
                return actividad

    except Exception as e:
        print(f"Ocurrió un error en def buscarActividadUnico: {e}")
        return []


def procesar_actualizar_actividad(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                codigo_actividad = data.form['codigo_actividad']
                nombre_actividad = data.form['nombre_actividad']
                descripcion_actividad = data.form['descripcion_actividad']
                id_actividad = data.form['id_actividad']             
                querySQL = """
                    UPDATE tbl_actividades
                    SET 
                        codigo_actividad = %s,
                        nombre_actividad = %s,
                        descripcion_actividad = %s
                    WHERE id_actividad = %s
                """
                values = (codigo_actividad, nombre_actividad, descripcion_actividad,id_actividad)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizar_actividad: {e}")
        return None


# Eliminar Actividades
def eliminarActividad(id_actividad):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_actividades WHERE id_actividad=%s"
                cursor.execute(querySQL, (id_actividad,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminaractividad : {e}")
        return []
    
    
    
    


### OPERACION DIARIA    
def obtener_id_empleados():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        DISTINCT e.id_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado ASC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
                
                # Extraer solo los valores de id_empleado de los diccionarios
                id_empleados = [empleado['id_empleado'] for empleado in empleadosBD]
        return id_empleados
    except Exception as e:
        print(f"Error en la función obtener_id_empleados: {e}")
        return None
    
def obtener_nombre_empleado(id_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                consulta = ("""SELECT CONCAT(nombre_empleado, ' ', apellido_empleado) as nombre_empleado FROM tbl_empleados WHERE id_empleado = %s""")
                cursor.execute(consulta, (id_empleado,))

                # Obtiene el resultado de la consulta
                resultado = cursor.fetchone()
                print(resultado)  # Imprime el resultado en la terminal

                # Retorna el nombre del empleado si hay un resultado
                return resultado

    except Exception as e:
        print(f"Error al obtener el nombre del empleado: {e}")
        return None

    finally:
        if 'cursor' in locals():
            cursor.close()
    
    
    
def obtener_proceso():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT DISTINCT
                        nombre_proceso  
                    FROM tbl_procesos AS e
                    ORDER BY e.nombre_proceso ASC
                    """)
                cursor.execute(querySQL,)
                procesosBD = cursor.fetchall()                
                # Extraer solo los valores de procesos de los diccionarios
                nombre_proceso = [proceso['nombre_proceso'] for proceso in procesosBD]
                print(nombre_proceso)
        return nombre_proceso
    except Exception as e:
        print(f"Error en la función obtener_nombre_proceso: {e}")
        return None
    
    
def obtener_actividad():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT DISTINCT
                        nombre_actividad
                    FROM tbl_actividades AS e
                    ORDER BY e.nombre_actividad ASC
                    """)
                cursor.execute(querySQL,)
                actividadBD = cursor.fetchall()                
                # Extraer solo los valores de actividad de los diccionarios
                nombre_actividad = [actividad['nombre_actividad'] for actividad in actividadBD]
                print(nombre_actividad)
        return nombre_actividad
    except Exception as e:
        print(f"Error en la función obtener_nombre_actividad: {e}")
        return None


def procesar_form_operacion(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO `tbl_operaciones` (`id_empleado`, `nombre_empleado`, `proceso`, `actividad`, `codigo_op`, `cantidad`, `novedad`, `fecha_hora_inicio`, `fecha_hora_fin`) VALUES  (%s, %s, %s,%s, %s, %s,%s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['id_empleado'], dataForm['nombre_empleado'], dataForm['nombre_proceso'], dataForm['nombre_actividad'], dataForm['cod_op'], dataForm['cantidad'], dataForm['novedades'], dataForm['hora_inicio'], dataForm['hora_fin'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_operacion: {str(e)}'
    
    
def sql_lista_operacionesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        o.id_operacion,
                        o.id_empleado,
                        o.nombre_empleado,
                        o.proceso,
                        o.actividad,
                        o.cantidad,
                        o.novedad,
                        o.fecha_hora_inicio,
                        o.fecha_hora_fin,
                        o.fecha_registro
                    FROM tbl_operaciones as o
                    ORDER BY fecha_registro DESC
                    """
                cursor.execute(querySQL)
                operacionesBD = cursor.fetchall()
        return operacionesBD
    except Exception as e:
        print(f"Error en la función sql_lista_operacionesBD: {e}")
        return None

def sql_detalles_operacionesBD(id_operacion):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        o.id_operacion,
                        o.id_empleado,
                        o.nombre_empleado,
                        o.proceso,
                        o.actividad,
                        o.codigo_op,
                        o.cantidad,
                        o.novedad,
                        o.fecha_hora_inicio,
                        o.fecha_hora_fin,
                        DATE_FORMAT(o.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_operaciones AS o
                    WHERE id_operacion =%s
                    """)
                cursor.execute(querySQL, (id_operacion,))
                operacionBD = cursor.fetchone()
        return operacionBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_operacionesBD: {e}")
        return None
    
def buscarOperacionUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            o.id_operacion,
                            o.id_empleado,
                            o.nombre_empleado,
                            o.proceso,
                            o.actividad,
                            o.cantidad,
                            o.novedad,
                            o.fecha_hora_inicio,
                            o.fecha_hora_fin,
                            o.fecha_registro
                        FROM tbl_operaciones AS o
                        WHERE o.id_operacion =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                operacion = mycursor.fetchone()
                return operacion

    except Exception as e:
        print(f"Ocurrió un error en def buscarOperacionUnico: {e}")
        return []
    
def procesar_actualizacion_operacion(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                id_operacion = data.form['id_operacion']
                proceso = data.form['proceso']
                actividad = data.form['actividad']
                cantidad = data.form['cantidad']
                novedad = data.form['novedad']             
                querySQL = """
                    UPDATE tbl_operaciones
                    SET 
                        proceso = %s,
                        actividad = %s,
                        cantidad = %s,
                        novedad = %s
                    WHERE id_operacion = %s
                """
                values = (proceso, actividad, cantidad,novedad,id_operacion)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizar_actividad: {e}")
        return None
    
# Eliminar OPeracion
def eliminarOperacion(id_operacion):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_operaciones WHERE id_operacion=%s"
                cursor.execute(querySQL, (id_operacion,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminar operacion : {e}")
        return []
    
    


### ORDEN DE PRODUCCION
def procesar_form_op(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_ordenProduccion (codigo_op, nombre_cliente, producto , estado, cantidad, odi , empleado) VALUES (%s, %s, %s, %s, %s, %s , %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['cod_op'], dataForm['nombre_cliente'], dataForm['producto'], dataForm['estado'], dataForm['cantidad'], dataForm['odi'], dataForm['vendedor'])
                print(valores)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_op: {str(e)}'


# Lista de Orden de Producción
def sql_lista_opBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        p.id_op,
                        p.codigo_op,
                        p.nombre_cliente,
                        p.producto,
                        p.estado,
                        p.cantidad, 
                        p.odi,
                        p.empleado,                     
                        p.fecha_registro
                    FROM tbl_ordenProduccion AS p
                    ORDER BY p.id_op DESC
                    """
                cursor.execute(querySQL)
                opBD = cursor.fetchall()
        return opBD
    except Exception as e:
        print(f"Error en la función sql_lista_opBD: {e}")
        return None


# Detalles del Orden de producción
def sql_detalles_opBD(idOp):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        p.id_op,
                        p.codigo_op,
                        p.nombre_cliente,
                        p.producto,
                        p.estado,
                        p.cantidad, 
                        p.odi,
                        p.empleado,  
                        DATE_FORMAT(P.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_ordenProduccion AS p
                    WHERE id_op =%s
                    ORDER BY p.id_op DESC
                    """)
                cursor.execute(querySQL, (idOp,))
                opBD = cursor.fetchone()
                print(opBD)
        return opBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_opBD: {e}")
        return None



def buscarOpUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            p.id_op,
                            p.codigo_op,
                            p.nombre_cliente,
                            p.producto,
                            p.estado,
                            p.cantidad, 
                            p.odi,
                            p.empleado,                        
                            p.fecha_registro
                        FROM tbl_ordenProduccion AS p
                        WHERE p.id_op =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                op = mycursor.fetchone()
                return op

    except Exception as e:
        print(f"Ocurrió un error en def buscarOpUnico: {e}")
        return []


def procesar_actualizar_form_op(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                codigo_op = data.form['codigo_op']
                nombre_cliente = data.form['nombre_cliente']
                producto = data.form['producto']
                estado = data.form['estado']
                cantidad = data.form['cantidad']
                odi = data.form['odi']
                empleado = data.form['empleado']
                id_op = data.form['id_op']             
                querySQL = """
                    UPDATE tbl_ordenProduccion
                    SET 
                        codigo_op = %s,
                        nombre_cliente = %s,
                        producto = %s,
                        estado = %s,
                        cantidad = %s,
                        odi = %s,
                        empleado = %s
                    WHERE id_op = %s
                """
                values = (codigo_op, nombre_cliente, producto,estado,cantidad,odi,empleado,id_op)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizar_form_op: {e}")
        return None

# Eliminar Orden de Producción
def eliminarOp(id_op):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_ordenProduccion WHERE id_op=%s"
                cursor.execute(querySQL, (id_op,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarOp : {e}")
        return []
    
def obtener_vendedor():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""SELECT DISTINCT CONCAT(nombre_empleado, ' ', apellido_empleado) as nombre_empleado FROM tbl_empleados order by nombre_empleado ASC""")
                cursor.execute(querySQL,)
                empleadoBD = cursor.fetchall()                
                # Extraer solo los valores de actividad de los diccionarios
                empleado = [empleado['nombre_empleado'] for empleado in empleadoBD]
        return empleado
    except Exception as e:
        print(f"Error en la función obtener_nombre_empleado: {e}")
        return None
    
def obtener_op():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT DISTINCT
                        o.codigo_op
                    FROM tbl_ordenproduccion AS o
                    ORDER BY o.codigo_op ASC
                    """)
                cursor.execute(querySQL,)
                opBD = cursor.fetchall()                
                # Extraer solo los valores de actividad de los diccionarios
                op = [op['codigo_op'] for op in opBD]
                print(op)
        return op
    except Exception as e:
        print(f"Error en la función obtener_nombre_op: {e}")
        return None
    
    
    


### JORNADA DIARIA    
def procesar_form_jornada(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO `tbl_jornadas` (`id_empleado`, `nombre_empleado`, `novedad_jornada_programada`, `novedad_jornada`, `fecha_hora_llegada_programada`, `fecha_hora_salida_programada`, `fecha_hora_llegada`, `fecha_hora_salida`) VALUES  (%s, %s, %s,%s, %s, %s,%s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['id_empleado'], dataForm['nombre_empleado'], dataForm['novedad_jornada_programada'], dataForm['novedad_jornada'], dataForm['fecha_hora_llegada_programada'], dataForm['fecha_hora_salida_programada'], dataForm['fecha_hora_llegada'], dataForm['fecha_hora_salida'])
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_jornada: {str(e)}'
    
    
def sql_lista_jornadasBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        o.id_jornada,
                        o.id_empleado,
                        o.nombre_empleado,
                        o.novedad_jornada_programada,
                        o.novedad_jornada,
                        o.fecha_hora_llegada_programada,
                        o.fecha_hora_salida_programada,
                        o.fecha_hora_llegada,
                        o.fecha_hora_salida,
                        o.fecha_registro
                    FROM tbl_jornadas as o
                    ORDER BY fecha_registro DESC
                    """
                cursor.execute(querySQL)
                jornadasBD = cursor.fetchall()
        return jornadasBD
    except Exception as e:
        print(f"Error en la función sql_lista_jornadasBD: {e}")
        return None

def sql_detalles_jornadasBD(id_jornada):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        o.id_jornada,
                        o.id_empleado,
                        o.nombre_empleado,
                        o.novedad_jornada_programada,
                        o.novedad_jornada,
                        o.fecha_hora_llegada_programada,
                        o.fecha_hora_salida_programada,
                        o.fecha_hora_llegada,
                        o.fecha_hora_salida,
                        DATE_FORMAT(o.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_jornadas AS o
                    WHERE id_jornada =%s
                    """)
                cursor.execute(querySQL, (id_jornada,))
                jornadaBD = cursor.fetchone()
        return jornadaBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_jornadasBD: {e}")
        return None
    
def buscarJornadaUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            o.id_jornada,
                            o.id_empleado,
                            o.nombre_empleado,
                            o.novedad_jornada_programada,
                            o.novedad_jornada,
                            o.fecha_hora_llegada_programada,
                            o.fecha_hora_salida_programada,
                            o.fecha_hora_llegada,
                            o.fecha_hora_salida,
                            o.fecha_registro
                        FROM tbl_jornadas AS o
                        WHERE o.id_jornada =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                jornada = mycursor.fetchone()
                return jornada

    except Exception as e:
        print(f"Ocurrió un error en def buscarjornadaUnico: {e}")
        return []
    
def procesar_actualizacion_jornada(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                id_jornada = data.form['id_jornada']
                id_empleado = data.form['id_empleado']
                nombre_empleado = data.form['nombre_empleado']
                novedad_jornada_programada = data.form['novedad_jornada_programada']
                novedad_jornada = data.form['novedad_jornada']
                fecha_hora_llegada_programada = data.form['fecha_hora_llegada_programada']
                fecha_hora_salida_programada = data.form['fecha_hora_salida_programada']  
                fecha_hora_llegada = data.form['fecha_hora_llegada']
                fecha_hora_salida = data.form['fecha_hora_salida']             
                querySQL = """
                    UPDATE tbl_jornadas
                    SET 
                        id_empleado = %s,
                        nombre_empleado = %s,
                        novedad_jornada_programada = %s,
                        novedad_jornada = %s,
                        fecha_hora_llegada_programada = %s,
                        fecha_hora_salida_programada = %s,
                        fecha_hora_llegada = %s,
                        fecha_hora_salida = %s
                    WHERE id_jornada = %s
                """
                values = (id_empleado, nombre_empleado, novedad_jornada_programada,novedad_jornada,fecha_hora_llegada_programada,fecha_hora_salida_programada,fecha_hora_llegada,fecha_hora_salida,id_jornada)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizar_jornada: {e}")
        return None
    
# Eliminar OPeracion
def eliminarJornada(id_jornada):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_jornadas WHERE id_jornada=%s"
                cursor.execute(querySQL, (id_jornada,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminar jornada : {e}")
        return []